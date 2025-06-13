import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from collections import defaultdict
from validation_logic import style_dataframe, safe_str, EXCEL_FIELDS

red_fill = PatternFill(start_color="FFFF6666", end_color="FFFF6666", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
orange_fill = PatternFill(start_color="FFFF9900", end_color="FFFF9900", fill_type="solid")

def clean_override_id(val):
    try:
        f = float(val)
        i = int(f)
        return str(i)
    except Exception:
        return str(val).strip()

def get_dynamic_suffixes(df):
    suffixes = set()
    for col in df.columns:
        if '_' in col:
            _, suffix = col.rsplit('_', 1)
            suffixes.add(suffix)
    return suffixes

def get_excel_and_csv_cols_for_suffix(df, suffix):
    excel_cols = [col for col in EXCEL_FIELDS if col in df.columns]
    csv_cols = [col for col in df.columns if col.endswith(f"_{suffix}")]
    return excel_cols, csv_cols

def run_validation_for_row(row, excel_cols, csv_cols, validate_cell_func, row_idx):
    all_valid = True
    for col in excel_cols + csv_cols:
        val = safe_str(row.get(col, ""))
        valid = validate_cell_func(col, val, row, row_idx)
        if valid is not True and valid != "csvgreen":
            all_valid = False
            break
    return all_valid

def process_files(excel_file, csv_files):
    try:
        excel_df = pd.read_excel(excel_file, dtype=str)
        excel_df["MFL ID"] = excel_df["MFL ID"].apply(clean_override_id)
        excel_df["OVERRIDE ID"] = excel_df["OVERRIDE ID"].apply(clean_override_id)
        excel_df["DATE TIME PRE KO (UTC)"] = pd.to_datetime(
            excel_df["DATE TIME PRE KO (UTC)"], errors='coerce', dayfirst=False, format="%Y-%m-%d %H:%M:%S"
        )

        main_merged = excel_df.copy()
        csv_data = {}
        unmatched_data = {}

        for csv_file in csv_files:
            label = csv_file.name.split('.')[0]
            df = pd.read_csv(csv_file, dtype=str)
            df["clientContentId"] = df["clientContentId"].apply(clean_override_id)
            df["performChannel"] = df["performChannel"].apply(clean_override_id)
            csv_data[label] = df

        inclusive_merged_rows = []
        used_csv_indexes = defaultdict(set)
        unmatched_excel_indexes = []

        for idx, excel_row in main_merged.iterrows():
            mfl_id = excel_row["MFL ID"]
            override_id = excel_row["OVERRIDE ID"]
            row_dict = excel_row.to_dict()

            matched_any = False

            for label, df in csv_data.items():
                suffix = f"_{label}"

                # Full match
                full_match = df[
                    (df["clientContentId"] == mfl_id) &
                    (df["performChannel"] == override_id)
                ]
                if not full_match.empty:
                    csv_row = full_match.iloc[0]
                    for c in df.columns:
                        row_dict[f"{c}{suffix}"] = csv_row[c]
                    row_dict[f"match_type{suffix}"] = "full"
                    used_csv_indexes[label].add(csv_row.name)
                    matched_any = True
                    continue

                # Partial match
                partial_match = df[
                    ((df["clientContentId"] == mfl_id) & (df["performChannel"] != override_id)) |
                    ((df["performChannel"] == override_id) & (df["clientContentId"] != mfl_id))
                ]
                if not partial_match.empty:
                    csv_row = partial_match.iloc[0]
                    for c in df.columns:
                        row_dict[f"{c}{suffix}"] = csv_row[c]
                    row_dict[f"match_type{suffix}"] = "partial"
                    row_dict[f"mismatch_key{suffix}"] = (
                        "MFL ID" if csv_row["performChannel"] == override_id else "OVERRIDE ID"
                    )
                    used_csv_indexes[label].add(csv_row.name)
                    matched_any = True
                else:
                    # Include keys as empty if not matched
                    for c in df.columns:
                        row_dict.setdefault(f"{c}{suffix}", "")
                    row_dict[f"match_type{suffix}"] = "none"

            if matched_any:
                inclusive_merged_rows.append(row_dict)
            else:
                unmatched_excel_indexes.append(idx)

        # Add unmatched Excel rows with empty CSV columns at the end
        for idx in unmatched_excel_indexes:
            excel_row = main_merged.loc[idx]
            row_dict = excel_row.to_dict()
            for label, df in csv_data.items():
                suffix = f"_{label}"
                for c in df.columns:
                    row_dict[f"{c}{suffix}"] = ""
                row_dict[f"match_type{suffix}"] = "none"
            inclusive_merged_rows.append(row_dict)

        # Ensure all key columns for each CSV are present, even if never matched
        if inclusive_merged_rows:
            for label, df in csv_data.items():
                suffix = f"_{label}"
                for c in df.columns:
                    colname = f"{c}{suffix}"
                    if colname not in inclusive_merged_rows[0]:
                        for row in inclusive_merged_rows:
                            row[colname] = ""

        merged_df = pd.DataFrame(inclusive_merged_rows)

        # --- New: Add validation to match_type columns ---
        # Get dynamic suffixes
        suffixes = get_dynamic_suffixes(merged_df)
        # Get validate_cell from your style_dataframe's closure
        validate_cell_func = style_dataframe.__globals__["validate_cell"]

        for suffix in suffixes:
            match_col = f"match_type_{suffix}"
            if match_col not in merged_df.columns:
                continue
            excel_cols, csv_cols = get_excel_and_csv_cols_for_suffix(merged_df, suffix)
            for idx, row in merged_df.iterrows():
                base_val = merged_df.at[idx, match_col]
                if base_val == "none":
                    continue  # Leave as "none"
                all_valid = run_validation_for_row(row, excel_cols, csv_cols, validate_cell_func, idx)
                merged_df.at[idx, match_col] = f"{base_val}+{'valid' if all_valid else 'invalid'}"

        # Remove match_date column from output if present (optional)
        if "match_date" in merged_df.columns:
            merged_df = merged_df.drop(columns=["match_date"])

        # Reorder columns: Excel columns first, then grouped CSV columns
        excel_cols = [col for col in excel_df.columns if col in merged_df.columns]
        csv_cols = [
            col for col in merged_df.columns
            if col not in excel_cols and not col.startswith("match_type") and not col.startswith("mismatch_key")
        ]
        field_groups = defaultdict(list)
        for col in csv_cols:
            base = col
            for label in csv_data.keys():
                suffix = f"_{label}"
                if col.endswith(suffix):
                    base = col[:-len(suffix)]
                    break
            field_groups[base].append(col)
        reordered_cols = excel_cols + [col for base in sorted(field_groups) for col in sorted(field_groups[base])]
        merged_df = merged_df[
            reordered_cols + [c for c in merged_df.columns if c.startswith("match_type") or c.startswith("mismatch_key")]
        ]

        # Output to Excel
        output = BytesIO()
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Merged Data"
        for r in dataframe_to_rows(merged_df, index=False, header=True):
            ws1.append(r)

        headers = [cell.value for cell in ws1[1]]
        # Highlight mismatches and partial matches
        for col_idx, col_name in enumerate(headers, start=1):
            if "_" in col_name:
                base_col = col_name.rsplit("_", 1)[0]
                label = col_name.rsplit("_", 1)[1]
                suffix = f"_{label}"
                match_type_col = f"match_type{suffix}"
                mismatch_key_col = f"mismatch_key{suffix}" if f"mismatch_key{suffix}" in headers else None

                if base_col in headers:
                    base_col_idx = headers.index(base_col) + 1
                    for row_idx in range(2, ws1.max_row + 1):
                        match_type = ws1.cell(row=row_idx, column=headers.index(match_type_col)+1).value if match_type_col in headers else ""
                        mismatch_key = ws1.cell(row=row_idx, column=headers.index(mismatch_key_col)+1).value if mismatch_key_col and mismatch_key_col in headers else ""
                        val = ws1.cell(row=row_idx, column=col_idx).value
                        base_val = ws1.cell(row=row_idx, column=base_col_idx).value

                        if match_type and "partial" in str(match_type) and mismatch_key and base_col == mismatch_key:
                            ws1.cell(row=row_idx, column=col_idx).fill = red_fill
                        elif str(val).strip() != str(base_val).strip() and match_type and "full" in str(match_type):
                            ws1.cell(row=row_idx, column=col_idx).fill = orange_fill
                        if val is None or str(val).strip() == "":
                            ws1.cell(row=row_idx, column=col_idx).fill = yellow_fill

        # Add unmatched CSV rows as separate sheets
        for label, df in csv_data.items():
            unmatched_indexes = set(df.index) - used_csv_indexes[label]
            unmatched = df.loc[list(unmatched_indexes)] if unmatched_indexes else pd.DataFrame(columns=df.columns)
            ws_csv = wb.create_sheet(f"Unmatched_{label[:25]}")
            if not unmatched.empty:
                for r in dataframe_to_rows(unmatched, index=False, header=True):
                    ws_csv.append(r)

        wb.save(output)
        output.seek(0)
        return {"detailed": output}

    except Exception as e:
        print("Error in process_files:", e)
        return None

def merge_files(excel_file, csv_files):
    output = process_files(excel_file, csv_files)
    return output, None, None

def save_output(merged_df, mismatch_summary, json_diff, filename):
    if merged_df is not None and hasattr(merged_df, "to_excel"):
        merged_df.to_excel(filename, index=False)
    pass