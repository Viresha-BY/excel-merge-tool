import pandas as pd
import json
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from collections import defaultdict

red_fill = PatternFill(start_color="FFFF6666", end_color="FFFF6666", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

def _is_invalid_key(val):
    try:
        return pd.isna(val) or str(val).strip().lower() in ("", "nan", "none")
    except Exception:
        return True

def clean_override_id(val):
    # Converts float strings like '1627.0' to '1627'
    try:
        f = float(val)
        i = int(f)
        if f == i:
            return str(i)
        else:
            return str(val).strip()
    except Exception:
        return str(val).strip()

def process_files(excel_file, csv_files, json_files):
    try:
        excel_df = pd.read_excel(excel_file)
        # Clean OVERRIDE ID to always be a string with no .0 decimal
        excel_df["MFL ID"] = excel_df["MFL ID"].apply(clean_override_id)
        excel_df["OVERRIDE ID"] = excel_df["OVERRIDE ID"].apply(clean_override_id)
        excel_df["DATE TIME PRE KO (UTC)"] = pd.to_datetime(excel_df["DATE TIME PRE KO (UTC)"], errors='coerce', dayfirst=True)
        excel_df["match_date"] = excel_df["DATE TIME PRE KO (UTC)"].dt.strftime("%Y-%m-%d")

        merged_df = excel_df.copy()
        csv_data = {}
        unmatched_data = {}

        for csv_file in csv_files:
            label = csv_file.name.split('.')[0]
            df = pd.read_csv(csv_file)
            df["clientContentId"] = df["clientContentId"].apply(clean_override_id)
            df.set_index("clientContentId", inplace=True, drop=False)
            csv_data[label] = df
            unmatched_data[label] = df.copy()

        # Batch-add all CSV columns for each label
        all_new_cols = []
        for label, df in csv_data.items():
            for col in df.columns:
                if col not in ("competitionId", "Day", "launchPeriod", "rightsId", "Source"):
                    all_new_cols.append(f"{col}_{label}")
        # Only add columns that don't already exist
        new_cols_to_add = [c for c in all_new_cols if c not in merged_df.columns]
        if new_cols_to_add:
            merged_df = pd.concat([merged_df, pd.DataFrame({col: None for col in new_cols_to_add}, index=merged_df.index)], axis=1)

        for i in merged_df.index:
            mfl_id = merged_df.at[i, "MFL ID"]
            if _is_invalid_key(mfl_id):
                continue
            for label, df in csv_data.items():
                if mfl_id in df.index:
                    row = df.loc[mfl_id]
                    if isinstance(row, pd.DataFrame):
                        row = row.iloc[0]
                    for col in df.columns:
                        if col not in ("competitionId", "Day", "launchPeriod", "rightsId", "Source"):
                            merged_df.at[i, f"{col}_{label}"] = row.get(col, None)
                    unmatched_data[label].drop(mfl_id, inplace=True, errors='ignore')

        # --- JSON section ---
        json_data_by_label = {}
        for json_file in json_files:
            label = json_file.name.split('.')[0]
            json_data = json.load(json_file)
            records = []
            for obj in json_data:
                e = obj.get("event", {})
                override_id_list = e.get("overrideId", [])
                override_id = (
                    clean_override_id(override_id_list[0].get("id"))
                    if override_id_list and isinstance(override_id_list[0], dict) and "id" in override_id_list[0]
                    else None
                )
                date_str = e.get("streamStartTime", "")[:10]
                oa_id = e.get("oaId", "")
                bcast = e.get("broadcasts", {}).get(oa_id, {})
                records.append({
                    "overrideId": override_id,
                    "date": date_str,
                    f"oaId_{label}": oa_id,
                    f"streamStartTime_{label}": e.get("streamStartTime", ""),
                    f"streamEndTime_{label}": e.get("streamEndTime", ""),
                    f"heEventTypeName_{label}": e.get("heEventTypeName", ""),
                    f"drmRequired_{label}": e.get("drmRequired", ""),
                    f"regions_{label}": ", ".join(e.get("regions", [])) if isinstance(e.get("regions", []), list) else e.get("regions", ""),
                    f"outputSuppressionMode_{label}": bcast.get("outputSuppressionMode", ""),
                    f"assetName_{label}": bcast.get("name", ""),
                    f"template_{label}": bcast.get("template", ""),
                    f"heResilience_{label}": e.get("heResilience", ""),
                    f"competitionId_{label}": e.get("competitionId", ""),
                    f"closedCaptioning_{label}": ", ".join(e.get("closedCaptioning", [])) if isinstance(e.get("closedCaptioning", []), list) else e.get("closedCaptioning", ""),
                    f"description_{label}": e.get("description", "")
                })
            json_df = pd.DataFrame(records).set_index(["overrideId", "date"])
            json_df = json_df.sort_index()  # Helps with MultiIndex performance
            json_data_by_label[label] = json_df

            # Batch-add all new JSON columns
            new_json_cols = [field for field in json_df.columns if field not in merged_df.columns]
            if new_json_cols:
                merged_df = pd.concat([merged_df, pd.DataFrame({col: None for col in new_json_cols}, index=merged_df.index)], axis=1)

            # --- Debugging block ---
            print(f"\n--- DEBUG for JSON label '{label}' ---")
            print("First 10 JSON keys:", list(json_df.index)[:10])
            print("First 10 Excel merge keys:",
                list(zip(merged_df["OVERRIDE ID"].astype(str).str.strip(), merged_df["match_date"].astype(str).str.strip()))[:10])

            match_count = 0
            for i in merged_df.index:
                override_id = merged_df.at[i, "OVERRIDE ID"]
                match_date = str(merged_df.at[i, "match_date"]).strip()
                # Skip nan/None/empty keys
                if _is_invalid_key(override_id) or _is_invalid_key(match_date):
                    if match_count < 10:
                        print(f"SKIPPING invalid key: ({override_id}, {match_date})")
                    continue
                key = (override_id, match_date)
                if match_count < 10:
                    print(f"Trying key: {key}")
                if key in json_df.index:
                    if match_count < 10:
                        print(f"JSON MATCH FOUND for key: {key}")
                    row = json_df.loc[key]
                    if isinstance(row, pd.DataFrame):
                        row = row.iloc[0]
                    for field in json_df.columns:
                        merged_df.at[i, field] = row.get(field, None)
                    match_count += 1

            print(f"Total JSON matches for '{label}': {match_count}")
            if match_count == 0:
                print(f"WARNING: No JSON data matched for '{label}'. Check your keys above for discrepancies!")

        merged_df.drop(columns=["match_date"], inplace=True)

        # Final column grouping by field name across labels
        all_labels = list(csv_data.keys()) + list(json_data_by_label.keys())
        excel_cols = [col for col in merged_df.columns if "_" not in col]
        grouped_cols = [col for col in merged_df.columns if "_" in col]

        field_groups = defaultdict(list)
        for col in grouped_cols:
            base, label = col.rsplit("_", 1)
            field_groups[base].append(col)

        reordered_cols = excel_cols + [col for base in sorted(field_groups) for col in sorted(field_groups[base])]
        merged_df = merged_df[reordered_cols]

        output = BytesIO()
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Merged Data"
        for r in dataframe_to_rows(merged_df, index=False, header=True):
            ws1.append(r)

        headers = [cell.value for cell in ws1[1]]
        for col_idx, col_name in enumerate(headers, start=1):
            if col_name.startswith("clientContentId_"):
                for row_idx in range(2, ws1.max_row + 1):
                    val = ws1.cell(row=row_idx, column=col_idx).value
                    mfl_val = ws1.cell(row=row_idx, column=headers.index("MFL ID") + 1).value
                    if val and mfl_val and str(val).strip() != str(mfl_val).strip():
                        ws1.cell(row=row_idx, column=col_idx).fill = red_fill
            elif "_" in col_name:
                for row_idx in range(2, ws1.max_row + 1):
                    val = ws1.cell(row=row_idx, column=col_idx).value
                    if val is None or str(val).strip() == "":
                        ws1.cell(row=row_idx, column=col_idx).fill = yellow_fill

        # Summary
        ws2 = wb.create_sheet("Summary")
        ws2.append(["File", "Type", "Total Rows", "Matched to Excel", "Extra", "Missing"])
        for label, df in csv_data.items():
            excel_ids = set(excel_df["MFL ID"])
            csv_ids = set(df["clientContentId"])
            matched = len(excel_ids & csv_ids)
            extra = len(csv_ids - excel_ids)
            missing = len(excel_ids - csv_ids)
            ws2.append([label, "CSV", len(csv_ids), matched, extra, missing])
        used_keys = set(zip(excel_df["OVERRIDE ID"], excel_df["DATE TIME PRE KO (UTC)"].dt.strftime("%Y-%m-%d")))
        for label, json_df in json_data_by_label.items():
            filtered = json_df.loc[json_df.index.isin(used_keys)]
            total = filtered.shape[0]
            filled = sum(filtered[field].notna().sum() for field in filtered.columns)
            empty = sum(filtered[field].isna().sum() for field in filtered.columns)
            ws2.append([label, "JSON", total, total, "-", empty])

        # Consolidated Summary
        ws3 = wb.create_sheet("Consolidated Summary")
        ws3.append(["Source", "File", "Field", "Matched", "Missing", "Mismatched"])
        for label, df in csv_data.items():
            for col in df.columns:
                if col not in ("competitionId", "Day", "launchPeriod", "rightsId", "Source"):
                    excel_col = col
                    merged_col = f"{col}_{label}"
                    matched = merged_df[merged_col].notna().sum()
                    missing = merged_df[merged_col].isna().sum()
                    mismatched = (merged_df[excel_col] != merged_df[merged_col]).sum()
                    ws3.append(["CSV", label, col, matched, missing, mismatched])
        for label, json_df in json_data_by_label.items():
            for col in json_df.columns:
                matched = merged_df[col].notna().sum()
                missing = merged_df[col].isna().sum()
                ws3.append(["JSON", label, col, matched, missing, "-"])

        wb.save(output)
        output.seek(0)
        return {"detailed": output, "clean": output}

    except Exception as e:
        print("Error in process_files:", e)
        return None

def merge_files(excel_file, csv_files, json_files):
    output = process_files(excel_file, csv_files, json_files)
    return output, None, None

def save_output(merged_df, mismatch_summary, json_diff, filename):
    if merged_df is not None and hasattr(merged_df, "to_excel"):
        merged_df.to_excel(filename, index=False)
    pass